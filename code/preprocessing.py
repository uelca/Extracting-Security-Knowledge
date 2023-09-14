import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os
import re


def read_file(txt_path):
    """
    This function takes the path to a .txt file in which
    the needed Bausteine from the IT-Grundschutzkompendium are placed.

    Returns:
    lines = []: A List of the Bausteine
    """
    lines = []
    with open(txt_path, 'r') as file:
        for line in file:
            # Remove leading and trailing whitespace or line breaks from the line
            line = line.strip()
            lines.append(line)
    return lines


def extract_table(excel_path, table_name):
    """
    This function takes the path to the Excel File which is called Kreuzreferenztabelle.
    In the Kreuzreferenztabelle file are Worksheets which represent each Baustein from the IT-Grundschutzkompendium.
    In this individual sheets the Threats and the Controls (which mitigate the threats) are saved.

    Furthermore it takes the name of a specific Worksheet.

    Returns:
    sheet: It returns the specific worksheet that is searched
    """
    workbook = load_workbook(excel_path)
    if table_name in workbook.sheetnames:
        # Create new Excel File which only holds the table of the sheet
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        new_workbook.create_sheet(title=table_name)
        for row in workbook[table_name].iter_rows(values_only=True):
            new_workbook.active.append(row)
        return new_workbook
    else:
        return None


def split_string(string, prefix, suffix):
    """
    This function takes a string, which is a Baustein from the IT-Grundschutzkompendium.
    However the Baustein from the IT-Grundschutzkompendium comes with a unique description like: OPS.1.2.3 and a name: Cloud Services
    To be able to execute the function extract_table it is needed to split the description of the Baustein at the space.
    Then it is needed to add the prefix "KRT_" and the suffix: ".xlsx". In this way it is possible to find the corresponding worksheet at the
    Kreuzreferenztabelle.

    The function could be also programmed to automatically add the predifined suffix and prefix, however they are kept as parameters to adjust for
    flexible use in future.

    Returns:
    sheet: It returns the string which is splitted at the first space and adds the suffix and prefix.
    """
    split_index = string.find(" ")
    if split_index != -1:
        string = prefix + string[:split_index] + suffix
    return string


def store_single_excel_tables(content_array, excel_path):
    for table_name in content_array:
        parent_folder = f'/Users/M.Fatih/PycharmProjects/Ontologie/Single Excel Tables/{table_name}'
        if not os.path.exists(parent_folder):
            os.makedirs(parent_folder)
        table_name = split_string(table_name, "KRT_", ".xlsx")

        # Extract the corresponding table in the krt2023_Excel.xlsx file for each line (component) in the Notwendige_Bausteine.txt file
        new_workbook = extract_table(excel_path, table_name)

        if new_workbook:
            # Name and save file by table name
            new_file_path = os.path.join(parent_folder, f'{table_name}')

            if os.path.exists(new_file_path):
                print(f"The file '{new_file_path}' already exists. The table will not be saved.")
            else:
                new_workbook.save(new_file_path)
                print(f"The table '{table_name}' has been saved in the file '{new_file_path}'.")
        else:
            print(f"The table {table_name} was not found in the Excel file.")


def create_text_files(root_folder):
    """
    Creates three text files in each sub-folder of the root folder.
    One text file is for the controls, the other one is for the threats and the third one is for the allocation of the conntrols to the threats

    :param root_folder: Path of the root folder
    """
    for root, dirs, files in os.walk(root_folder):
        for dir_name in dirs:
            # Path to the current subdirectory
            subdir_path = os.path.join(root, dir_name)

            # Create the name for the controls file
            controls_file_name = f"{dir_name}-Controls.txt"
            controls_file_path = os.path.join(subdir_path, controls_file_name)

            # Check if the controls file already exists
            if not os.path.exists(controls_file_path):
                # Create the controls file
                try:
                    with open(controls_file_path, 'w') as controls_file:
                        pass  # No action required
                    print(f"Controls file created: {controls_file_path}")
                except Exception as e:
                    print(f"Error creating controls file: {controls_file_path}")
                    print(f"Error message: {str(e)}")

            # Create the name for the threats file
            threats_file_name = f"{dir_name}-Threats.txt"
            threats_file_path = os.path.join(subdir_path, threats_file_name)

            # Check if the threats file already exists
            if not os.path.exists(threats_file_path):
                # Create the threats file
                try:
                    with open(threats_file_path, 'w') as threats_file:
                        pass  # No action required
                    print(f"Threats file created: {threats_file_path}")
                except Exception as e:
                    print(f"Error creating threats file: {threats_file_path}")
                    print(f"Error message: {str(e)}")

            # Create the name for the Allocation file
            allocation_file_name = f"{dir_name}-Allocation.txt"
            allocation_file_path = os.path.join(subdir_path, allocation_file_name)

            # Check if the threats file already exists
            if not os.path.exists(allocation_file_path):
                # Create the threats file
                try:
                    with open(allocation_file_path, 'w') as allocation_file:
                        pass  # No action required
                    print(f"Allocations file created: {allocation_file_path}")
                except Exception as e:
                    print(f"Error creating allocation file: {allocation_file_path}")
                    print(f"Error message: {str(e)}")


def process_excel_files_in_folder(root_directory):
    """
    In this Method all subfolders of the root directory are iterated. Each subfolder has following structure:
    An Excel File with the stored Reference Table(KRT) for the corresponding Baustein.
    Two Text Files which are used to store the Controls and Threats from the KRT.
    :param root_directory: Directory of the Folder which contains all subfolders
    """
    # Iterate over all subfolders and access files
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file.endswith(".xlsx"):
                # Determine the paths of the corresponding files
                excel_file_path = os.path.join(root, file)
            elif file.endswith("-Controls.txt"):
                controls_file_path = os.path.join(root, file)
            elif file.endswith("-Threats.txt"):
                threats_file_path = os.path.join(root, file)
            elif file.endswith("-Allocation.txt"):
                allocations_file_path = os.path.join(root, file)
        if files:
            process_excel_file(excel_file_path, controls_file_path, threats_file_path, allocations_file_path)


def process_excel_file(excel_file_path, controls_file_path, threats_file_path, allocations_file_path):
    """
    this method is executed for each subfolder in the Single Excel Tables Folder.
    Basically the single Excel File (KRT) is taken and then each Control and each Threat are extracted and written to the
    corresponding .txt file.
    Additionaly in the allocations file is saved. This file stores which control mitigates which threat.

    :param excel_file_path: The file path to the single excel file (KRT)
    :param controls_file_path: The file path to the controls text file
    :param threats_file_path: The file path to the threats text file
    :param allocations_file_path: The file path to the allocations text file
    :return:
    """
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the active worksheet
    worksheet = workbook.active

    # File paths for rows and columns text files
    controls_text_file = controls_file_path
    threats_text_file = threats_file_path
    allocations_text_file = allocations_file_path

    # Write Controls and the allocation: "Control-Threat" to the respective text files
    with open(controls_text_file, 'w') as controls_file, open(allocations_text_file, 'w') as allocations_file:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] == "ENTFALLEN":
                continue

            # While iterating over the rows (Controls) we simultaneously iterate over each cell to store which threat is mitigated by the control
            # To do so we save the indices where a "X" is placed
            x_indices = []
            for index, cell in enumerate(row[3:], start=3):  # Start iterating from the third cell
                if cell == "X":
                    x_indices.append(index)

            if not x_indices:
                continue

            # Get the column letters from the indices
            column_letters = []
            for x_index in x_indices:
                # Calculate the column index for the corresponding x_index
                column_index = ""
                while x_index >= 0:
                    remainder = x_index % 26
                    column_index = chr(ord('A') + remainder) + column_index
                    x_index = (x_index - remainder) // 26 - 1 if x_index > 0 else -1

                column_letters.append(column_index)

            # Get the values of the first cells in the columns
            first_cell_values = [worksheet[f"{column_letter}1"].value for column_letter in column_letters]
            # Write Controls to a text file
            row_data = ' '.join(str(cell) for index, cell in enumerate(row) if index <= 1)
            row_data.rstrip()
            controls_file.write(row_data + ":" + '\n' +
                                "Description: " + find_description("../data/XML_Kompendium_2023.xml", row_data) + '\n\n')
            # Write Allocations to a text file
            for value in first_cell_values:
                allocations_file.write(row_data + ' ' + "mitigates" + ' ' + str(value) + '\n\n')

    # Write Threats to a text file
    with open(threats_text_file, 'w') as threats:
        for column in worksheet.iter_cols(min_col=4, values_only=True):
            column_data = ' '.join(str(cell) for index, cell in enumerate(column) if index < 1)
            # For Threats, we additionally need to get the Threat Title
            threats.write(find_threat_title("../data/XML_Kompendium_2023.xml", column_data) + ":" + '\n' +
                          "Description: " + find_description("../data/XML_Kompendium_2023.xml", column_data) + '\n\n')

    # Close the Excel file
    workbook.close()


def find_description(xml_file_path, search_text):
    """
    In this Method we are searching for the description of the Control or the Threat. This is needed, since in the
    next step we need to input it to ChatGPT
    :param xml_file_path: Path to the XML Version of the grundschutzkompendium
    :param search_text: The Control for which we are searching
    :return: The Description of the Control or Threat
    """
    try:
        with open(xml_file_path, 'r') as xml_file:
            xml_data = xml_file.read()

        # Search for the search text in the XML
        match = re.search(search_text, xml_data)

        if match:
            # Extract the text between <para> and </para> on the following line
            para_match = re.search(r'<para>(.*?)</para>', xml_data[match.end():], re.DOTALL)
            if para_match:
                extracted_text = para_match.group(1)
                return extracted_text.strip()  # Remove leading and trailing whitespace and line breaks
        else:
            # Search for the whole Description was not successful, now just search with Identifier, e.g. INF.1.1.A1
            split_string = search_text.split(" ")
            identifier = split_string[0]
            search_result = re.findall(split_string[0], xml_data)
            # If there is a unique result for that identifier, we can use the description part in next line
            if len(search_result) == 1:
                second_match = re.search(identifier, xml_data)
                second_para_match = re.search(r'<para>(.*?)</para>', xml_data[second_match.end():], re.DOTALL)
                if second_para_match:
                    extracted_text = second_para_match.group(1)
                    return extracted_text.strip()  # Remove leading and trailing whitespace and line breaks
            else:
                print("Not unique, please fill " + identifier + " Description by your own")
                return "Search text not found."

    except FileNotFoundError:
        print(search_text + " not found")
        return "File not found."


def find_threat_title(xml_file_path, search_text):
    try:
        with open(xml_file_path, 'r') as xml_file:
            xml_data = xml_file.read()

        # Suchen nach dem Suchtext im XML
        match = re.search(search_text, xml_data)

        if match:
            # Extrahiere die gesamte Zeile, in der der Suchtext gefunden wurde
            start = xml_data.rfind('\n', 0, match.start()) + 1
            end = xml_data.find('\n', match.end())
            extracted_line = xml_data[start:end].strip()  # Bereinige die Zeile
            extracted_line = extracted_line[len("<title>"): -len("</title>")]
            return extracted_line
        else:
            print(search_text + " not found")
            return "Search title not found."

    except FileNotFoundError:
        print(search_text + " not found")
        return "File not found."

def preprocess_data(txt_path, excel_path, root_folder):
    # content_array = read_file(txt_path)
    # store_single_excel_tables(content_array, excel_path)
    # create_text_files(root_folder)
    process_excel_files_in_folder(root_folder)