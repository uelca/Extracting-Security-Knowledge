import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os


def read_file(txt_path):
    """
    This function takes the path to a .txt file in which
    the needed Bausteine from the IT-Grundschutzkompendium are placed.

    Returns:
    lines = []: A List of the Bausteine
    """
    lines = []
    with open(txt_path, 'r') as file:
        for line in file:
            # Remove leading and trailing whitespace or line breaks from the line
            line = line.strip()
            lines.append(line)
    return lines


def extract_table(excel_path, table_name):
    """
    This function takes the path to Excel File which is called Kreuzreferenztabelle
    In this Excel file are Worksheets which represent each Baustein from the IT-Grundschutzkompendium.

    Furthermore it takes the name of a specific Worksheet.

    Returns:
    sheet: It returns the specific worksheet that is searched
    """
    workbook = load_workbook(excel_path)
    if table_name in workbook.sheetnames:
        # Create new Excel File which only holds the table of the sheet
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        new_workbook.create_sheet(title=table_name)
        for row in workbook[table_name].iter_rows(values_only=True):
            new_workbook.active.append(row)
        return new_workbook
    else:
        return None


def split_string(string, prefix, suffix):
    """
    This function takes a string, which is a Baustein from the IT-Grundschutzkompendium.
    However the Baustein from the IT-Grundschutzkompendium comes with a unique description like: OPS.1.2.3 and a name: Cloud Services
    To be able to execute the function extract_table it is needed to split the description of the Baustein at the space.
    Then it is needed to add the prefix "KRT_" and the suffix: ".xlsx". In this way it is possible to find the corresponding worksheet at the
    Kreuzreferenztabelle.

    The function could be also programmed to automatically add the predifined suffix and prefix, however they are kept as parameters to adjust for
    flexible use in future.

    Returns:
    sheet: It returns the string which is splitted at the first space and adds the suffix and prefix.
    """
    split_index = string.find(" ")
    if split_index != -1:
        string = prefix + string[:split_index] + suffix
    return string


def store_single_excel_tables(content_array, excel_path):
    for table_name in content_array:
        parent_folder = f'/Users/M.Fatih/PycharmProjects/Ontologie/Single Excel Tables/{table_name}'
        if not os.path.exists(parent_folder):
            os.makedirs(parent_folder)
        table_name = split_string(table_name, "KRT_", ".xlsx")

        # Extract the corresponding table in the krt2023_Excel.xlsx file for each line (component) in the Notwendige_Bausteine.txt file
        new_workbook = extract_table(excel_path, table_name)

        if new_workbook:
            # Name and save file by table name
            new_file_path = os.path.join(parent_folder, f'{table_name}')

            if os.path.exists(new_file_path):
                print(f"The file '{new_file_path}' already exists. The table will not be saved.")
            else:
                new_workbook.save(new_file_path)
                print(f"The table '{table_name}' has been saved in the file '{new_file_path}'.")
        else:
            print(f"The table {table_name} was not found in the Excel file.")


def create_text_files(root_folder):
    """
    Creates three text files in each sub-folder of the root folder.
    One text file is for the controls, te other one is for the threats and the third one is for the allocation of the conntrols to the threats

    :param root_folder: Path of the root folder
    """
    for root, dirs, files in os.walk(root_folder):
        for dir_name in dirs:
            # Path to the current subdirectory
            subdir_path = os.path.join(root, dir_name)

            # Create the name for the controls file
            controls_file_name = f"{dir_name}-Controls.txt"
            controls_file_path = os.path.join(subdir_path, controls_file_name)

            # Check if the controls file already exists
            if not os.path.exists(controls_file_path):
                # Create the controls file
                try:
                    with open(controls_file_path, 'w') as controls_file:
                        pass  # No action required
                    print(f"Controls file created: {controls_file_path}")
                except Exception as e:
                    print(f"Error creating controls file: {controls_file_path}")
                    print(f"Error message: {str(e)}")

            # Create the name for the threats file
            threats_file_name = f"{dir_name}-Threats.txt"
            threats_file_path = os.path.join(subdir_path, threats_file_name)

            # Check if the threats file already exists
            if not os.path.exists(threats_file_path):
                # Create the threats file
                try:
                    with open(threats_file_path, 'w') as threats_file:
                        pass  # No action required
                    print(f"Threats file created: {threats_file_path}")
                except Exception as e:
                    print(f"Error creating threats file: {threats_file_path}")
                    print(f"Error message: {str(e)}")

            # Create the name for the Allocation file
            allocation_file_name = f"{dir_name}-Allocation.txt"
            allocation_file_path = os.path.join(subdir_path, allocation_file_name)

            # Check if the threats file already exists
            if not os.path.exists(allocation_file_path):
                # Create the threats file
                try:
                    with open(allocation_file_path, 'w') as allocation_file:
                        pass  # No action required
                    print(f"Allocations file created: {allocation_file_path}")
                except Exception as e:
                    print(f"Error creating allocation file: {allocation_file_path}")
                    print(f"Error message: {str(e)}")

def process_excel_file(excel_file_path, controls_file_path, threats_file_path, allocations_file_path):
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the active worksheet
    worksheet = workbook.active

    # File paths for rows and columns text files
    controls_text_file = controls_file_path
    threats_text_file = threats_file_path
    allocations_text_file = allocations_file_path


    # Write Controls and the allocation: "Control-Threat" to the respective text files
    with open(controls_text_file, 'w') as controls_file, open(allocations_text_file, 'w') as allocations_file:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] == "ENTFALLEN":
                continue

            # While iterating over the rows (Controls) we simultaneously iterate over each cell to store which threat is mitigated by the control
            # To do so we save the indices where a "X" is placed
            x_indices = []
            for index, cell in enumerate(row[2:], start=2):  # Start iterating from the third cell
                if cell == "X":
                    x_indices.append(index)

            if not x_indices:
                continue

            # Get the column letters from the indices
            column_letters = [chr(ord('A') + x_index) for x_index in x_indices]

            # Get the values of the first cells in the columns
            first_cell_values = [worksheet[f"{column_letter}1"].value for column_letter in column_letters]

            row_data = '\t'.join(str(cell) for index, cell in enumerate(row) if index <= 1)
            controls_file.write(row_data + '\n')
            for value in first_cell_values:
                allocations_file.write(row_data + '\t' + str(value) + '\n')

    # Write Threats to a text file
    with open(threats_text_file, 'w') as file:
        for column in worksheet.iter_cols(min_col=4, values_only=True):
            column_data = '\t'.join(str(cell) for index, cell in enumerate(column) if index < 1)
            file.write(column_data + '\n')

    # Close the Excel file
    workbook.close()


def process_excel_files_in_folder(root_directory):
    """
    In this Method all subfolders of the root directory are iterated. Each subfolder has following structure:
    An Excel File with the stored Reference Table(KRT) for the corresponding Baustein.
    Two Text Files which are used to store the Controls and Threats from the KRT.
    :param root_directory: Directory of the Folder which contains all subfolders
    """
    # Iterate over all subfolders and access files
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file.endswith(".xlsx"):
                # Determine the paths of the corresponding files
                excel_file_path = os.path.join(root, file)
            elif file.endswith("-Controls.txt"):
                controls_file_path = os.path.join(root, file)
            elif file.endswith("-Threats.txt"):
                threats_file_path = os.path.join(root, file)
            elif file.endswith("-Allocation.txt"):
                allocations_file_path = os.path.join(root, file)
        if files:
            process_excel_file(excel_file_path, controls_file_path, threats_file_path, allocations_file_path)


if __name__ == '__main__':
    # Path to the text file with Bausteine
    txt_path = '/Users/M.Fatih/PycharmProjects/Ontologie/Notwendige Bausteine.txt'
    excel_path = '/Users/M.Fatih/PycharmProjects/Ontologie/krt2023_Excel.xlsx'
    root_folder = '/Users/M.Fatih/PycharmProjects/Ontologie/Single Excel Tables'

    # Create an array to store the contents of the Notwendige_Bausteine.txt file
    # content_array = read_file(txt_path)
    # store_single_excel_tables(content_array, excel_path)
    # create_text_files(root_folder)
    process_excel_files_in_folder(root_folder)
